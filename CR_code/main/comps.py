# -*- coding: utf-8 -*-
"""
Created on Mon Jun 13 19:06:05 2022

@author: eduardo.scheffer
"""

''' 1) IMPORT MAIN LIBRARIES --------------------------------------------------------------------------------''' 
import pandas as pd
import numpy as np
import math as math

import datetime as dt

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

#import matplotlib.pyplot as plt
#import seaborn as sns

import warnings 
warnings.filterwarnings('ignore')

pd.set_option("display.max_columns", None)
pd.set_option("display.max_rows", 20)

import sys
import os
import inspect

parent_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))))
sys.path.append(parent_path+'/Comps_code/formulas')


from  fund_prices_database import fund_prices_database
from  index_prices_database import index_prices_database
from drawdown import event_drawdown, max_drawdown
from moving_window import moving_window
#from  fund_prices_cvm import fund_prices_cvm

''' 2) READ SELECTED FUNDS ------------------------------------------------------------------------------------''' 

fund_universe = "Funds List"

excel_path = parent_path + '/Comparables.xlsx'

funds_list = pd.read_excel(excel_path, sheet_name = fund_universe).iloc[2:,1:6].dropna(how='all',axis='columns')
funds_list = funds_list[(funds_list.iloc[:,3]!="")]
funds_list = funds_list.drop(funds_list.columns[1], axis=1)
funds_name = list(funds_list.iloc[:,0:1].to_numpy().flatten())
cnpj_list = list(funds_list.iloc[:,1:2].to_numpy().flatten().astype(float))
cnpj_class_strategy = funds_list.set_index(funds_list.columns[0]).sort_index(axis=0)
cnpj_class_strategy.columns= ['CNPJ', 'Classe', 'Estratégia']
cnpj_class_strategy.index.names = ['index']

dict_map = dict(zip(cnpj_list, funds_name))

#cnpj_list = [str(x) for x in cnpj_list] 
#for i in range(len(cnpj_list)):
#    if len(cnpj_list[i]) == 13:
#        cnpj_list[i] = '0'+ cnpj_list[i]
        
''' 3) IMPORT FUND AND BENCHMARK PRICES --------------------------------------------------------------------------------------'''

benchmark_list = list(["CDI", "IHFA", "Ibovespa"])
   
benchmark = index_prices_database(benchmark_list)    

price, fund_age, fund_start = fund_prices_database(cnpj_list) 

# Change CNPJ for fund name
funds_columns = list(price.columns)
funds_columns = [dict_map.get(item, item) for item in funds_columns] #susbtitute cnpj by the name from the excel sheet
price.columns = funds_columns

fund_age_names = list(fund_age.index)
fund_age_names = [dict_map.get(item, item) for item in fund_age_names] #susbtitute cnpj by the name from the excel sheet
fund_age = fund_age.set_index([fund_age_names])

fund_start_names = list(fund_start.index)
fund_start_names = [dict_map.get(item, item) for item in fund_start_names] #susbtitute cnpj by the name from the excel sheet
fund_start = fund_start.set_index([fund_start_names])

# Organize alphabetically
fund_age = fund_age.sort_index(axis=0)
fund_start = fund_start.sort_index(axis=0)

# Concat benchmarks and funds
fund_age = pd.concat([pd.DataFrame(columns = ['Age (months)'], index = benchmark_list), fund_age])
fund_start = pd.concat([pd.DataFrame(columns = ['Dates'], index = benchmark_list), fund_start])


#Filter funds that have data
cnpj_class_strategy = cnpj_class_strategy.reset_index()
aux = fund_age.reset_index()
cnpj_class_strategy = pd.merge(cnpj_class_strategy,aux, on='index', how='right' ).drop(["Age (months)"], axis=1).drop_duplicates()

cnpj_class_strategy = cnpj_class_strategy.set_index(cnpj_class_strategy.columns[0])
                            
''' 4) CALCULATE LN RETURNS --------------------------------------------------------------------------------------'''

fund_lnReturns = np.log(price.astype('float')) - np.log(price.astype('float').shift(1))
fund_lnReturns = fund_lnReturns.reindex(sorted(fund_lnReturns.columns), axis=1) # Organize alphabetically by fund name

benchmark_lnReturns = np.log(benchmark.astype('float')) - np.log(benchmark.astype('float').shift(1))
benchmark_lnReturns.loc[:,'CDI'] = (1+benchmark.loc[:,'CDI'].astype('float')/100)**(1/252)-1

ln_Returns = pd.merge(benchmark_lnReturns, fund_lnReturns, right_index=True, left_index=True)
ln_Returns = ln_Returns.iloc[1:,:]


dates = pd.DataFrame(ln_Returns.index[:].to_pydatetime(), columns=["Dates"])

# get row that begin each fund

fund_start['Dates'] = pd.to_datetime(fund_start['Dates'])
for i in funds_columns:
    fund_start.loc[i, 'Dates'] = dates[dates['Dates']>=fund_start.loc[i, 'Dates']].iloc[0,0]
fund_start = fund_start.reset_index()
fund_start = fund_start.rename(columns={fund_start.columns[0]: 'Nome'})
dates_rows = dates.reset_index()
fund_start = pd.merge(fund_start,dates_rows, on='Dates', how='left').fillna(0)
fund_start.set_index('Nome', inplace=True)


# get last date
date_last = dates.iloc[-1, 0]
M_last = date_last.month
Y_last = date_last.year

# Get start dates for each period
date_MTD = min(dates[(dates['Dates'] > dt.datetime(Y_last, M_last, 1))]['Dates']) # get first return date for MTD
date_YTD = min(dates[(dates['Dates'] > dt.datetime(Y_last, 1, 1))]['Dates']) # get first return date for YTD
date_3M = date_last - dt.timedelta(days=91) # get first return date for 3M
date_6M = date_last - dt.timedelta(days=182)
date_12M = date_last - dt.timedelta(days=365)
date_24M = date_last - dt.timedelta(days=730)
date_36M = date_last - dt.timedelta(days=3*365)


#Delete funds that begun in 2022
#ln_Returns = ln_Returns.loc[:, (ln_Returns[(ln_Returns.index >= date_YTD)].iloc[0:3].isna().sum()<2)]

'''5) CALCULATE PERFORMANCE INDICATORS ----------------------------------------------------------------------------'''

print('Performance indicators calculating...')

summary_cols = ['CNPJ', 'Classe', 'Estratégia', 'Ret_MTD', 'Ret_YTD', 'Ret_12M', 'Ret_24M', 'Ret_36M',
                'Sharpe_YTD', 'Sharpe_12M', 'Sharpe_24M', 'Sharpe_36M',
                'Vol_12M', 'Vol_24M', 'Beta',
                'DD_JD', 'DD_GC', "DD_Covid", 'DD_Max', 
                'TTR_JD', 'TTR_GC', "TTR_Covid", 'TTR_Max', 
                'Histórico (meses)', '% meses acima do CDI',  
                'Ret_moving_3M_median', 'Ret_moving_3M_min', 'Ret_moving_3M_max', 
                'Ret_moving_6M_median', 'Ret_moving_6M_min', 'Ret_moving_6M_max',
                'Ret_moving_12M_median', 'Ret_moving_12M_min', 'Ret_moving_12M_max',
                'Sharp_moving_12M_median', 'Sharp_moving_12M_min', 'Sharp_moving_12M_max', 
                'Consistency_24M', 'Consistency_36M']

summary_rows = list(ln_Returns.columns)
summary = pd.DataFrame(columns = summary_cols, index = summary_rows)

# CNPJ, Class and Strategy
summary['CNPJ'] = cnpj_class_strategy.iloc[:,0]
summary['Classe'] = cnpj_class_strategy.iloc[:,1]
summary['Estratégia'] = cnpj_class_strategy.iloc[:,2]

# Get age of each fund
summary['Histórico (meses)'] = list(fund_age.iloc[:,0])

# Get returns for different periods
summary['Ret_MTD'] = ln_Returns[(ln_Returns.index >= date_MTD) & (ln_Returns.index <= date_last)].sum()
summary['Ret_YTD'] = ln_Returns[(ln_Returns.index >= date_YTD) & (ln_Returns.index <= date_last)].sum()
summary['Ret_12M'] = ln_Returns[(ln_Returns.index >= date_12M) & (ln_Returns.index <= date_last)].sum()
summary['Ret_24M'] = ln_Returns[(ln_Returns.index >= date_24M) & (ln_Returns.index <= date_last)].sum()
summary['Ret_36M'] = ln_Returns[(ln_Returns.index >= date_36M) & (ln_Returns.index <= date_last)].sum()

# Get volatility for different periods
vol_YTD = ln_Returns[(ln_Returns.index >= date_YTD) & (ln_Returns.index <= date_last)].std()*np.sqrt(252)
summary['Vol_12M'] = ln_Returns[(ln_Returns.index >= date_12M) & (ln_Returns.index <= date_last)].std()*np.sqrt(252)
summary['Vol_24M'] = ln_Returns[(ln_Returns.index >= date_24M) & (ln_Returns.index <= date_last)].std()*np.sqrt(252)
vol_36M = ln_Returns[(ln_Returns.index >= date_36M) & (ln_Returns.index <= date_last)].std()*np.sqrt(252)

# Get sharpe for different periods
days_YTD = ln_Returns[(ln_Returns.index >= date_YTD) & (ln_Returns.index <= date_last)].count()[0]
days_12M = ln_Returns[(ln_Returns.index >= date_12M) & (ln_Returns.index <= date_last)].count()[0]
days_24M = ln_Returns[(ln_Returns.index >= date_24M) & (ln_Returns.index <= date_last)].count()[0]
days_36M = ln_Returns[(ln_Returns.index >= date_36M) & (ln_Returns.index <= date_last)].count()[0]
summary['Sharpe_YTD'] = ((1+summary['Ret_YTD'])**(252/days_YTD) - np.ones(len(summary))*(1+summary.loc['CDI', 'Ret_YTD'])**(252/days_YTD))/vol_YTD
summary['Sharpe_12M'] = ((1+summary['Ret_12M'])**(252/days_12M) - np.ones(len(summary))*(1+summary.loc['CDI', 'Ret_YTD'])**(252/days_12M))/summary['Vol_12M']
summary['Sharpe_24M'] = ((1+summary['Ret_24M'])**(252/days_24M) - np.ones(len(summary))*(1+summary.loc['CDI', 'Ret_YTD'])**(252/days_24M))/summary['Vol_24M']
summary['Sharpe_36M'] = ((1+summary['Ret_24M'])**(252/days_36M) - np.ones(len(summary))*(1+summary.loc['CDI', 'Ret_YTD'])**(252/days_36M))/vol_36M



# Get % of months above CDI
M_i = dates.iloc[0,0].month
Y_i = dates.iloc[0,0].year
M_last = date_last.month - 1
if M_last == 0:
    M_last = 12
    Y_last = date_last.year - 1
else: 
    Y_last = date_last.year

summary['% meses acima do CDI'] =np.zeros(len(summary['% meses acima do CDI'])).astype(int)
n_months = np.zeros(len(summary['% meses acima do CDI'])).astype(int)
while (M_i != M_last) or (Y_i != Y_last):
    ret_M = ln_Returns[(ln_Returns.index.month == M_i) & (ln_Returns.index.year == Y_i)].sum()
    ret_CDI_M = ln_Returns[(ln_Returns.index.month == M_i) & (ln_Returns.index.year == Y_i)]['CDI'].sum()
    summary['% meses acima do CDI'] = summary['% meses acima do CDI'] + (ret_M > ret_CDI_M)*1
    n_months = n_months + (ret_M != 0)*1
    if M_i <= 11:
        M_i = M_i + 1
    else:
        M_i = 1
        Y_i = Y_i + 1

summary['% meses acima do CDI'] = summary['% meses acima do CDI']/n_months    

# Moving window statistics using 36 months
step = 10
fund_start_after_36 = fund_start.copy()
fund_start_after_36['index'] = fund_start_after_36['index'] - (len(dates) - days_36M)
fund_start_after_36 = fund_start_after_36[fund_start_after_36['index']>0] # Get funds that start after the window starts
fund_start_after_36['index'] = fund_start_after_36['index'].apply(lambda x: 0 if x<0 else math.ceil(x/step)) # Get the windownumber in which those funds start

summary['Consistency_36M'], persistency_perc = moving_window('consistency', ln_Returns[(ln_Returns.index >= date_36M)], step, 252, fund_start_after_36)

# Moving window statistics using 24 months
step = 10
fund_start_24 = fund_start.copy()
fund_start_24['index'] = fund_start_24['index'] - (len(dates) - days_24M)
fund_start_24 = fund_start_24[fund_start_24['index']>0] # Get funds that start after the window starts
fund_start_24['index'] = fund_start_24['index'].apply(lambda x: 0 if x<0 else math.ceil(x/step)) # Get the windownumber in which those funds start

summary['Ret_moving_3M_median'], summary['Ret_moving_3M_min'], summary['Ret_moving_3M_max'] = moving_window('return', ln_Returns[(ln_Returns.index >= date_24M)], step, 3*21, fund_start_24)
summary['Ret_moving_6M_median'], summary['Ret_moving_6M_min'], summary['Ret_moving_6M_max'] = moving_window('return', ln_Returns[(ln_Returns.index >= date_24M)], step, 6*21, fund_start_24)
summary['Ret_moving_12M_median'], summary['Ret_moving_12M_min'], summary['Ret_moving_12M_max'] = moving_window('return', ln_Returns[(ln_Returns.index >= date_24M)], step, 252, fund_start_24)
summary['Sharp_moving_12M_median'], summary['Sharp_moving_12M_min'], summary['Sharp_moving_12M_max'] = moving_window('sharpe', ln_Returns[(ln_Returns.index >= date_24M)], step, 252, fund_start_24)
summary['Consistency_24M'], persistency_perc = moving_window('consistency', ln_Returns[(ln_Returns.index >= date_24M)], step, 252, fund_start_24)



# Get drawdowns in stress periods
print('Drawdowns calculating...')

JD_date = dt.datetime(2017, 5, 17).strftime("%Y-%m-%d") # Joesley Day
GC_date = dt.datetime(2018, 5, 17).strftime("%Y-%m-%d") # Greve dos Caminhoneiros
Covid_date = dt.datetime(2020, 2, 26).strftime("%Y-%m-%d") # Covid
Max_date = dates.iloc[0,0].strftime("%Y-%m-%d") # All period
ln_Returns = ln_Returns.fillna(0)

#summary['DD_JD'], summary['TTR_JD'] = event_drawdown(ln_Returns[(ln_Returns.index >= JD_date)], 5) 
#summary['DD_GC'], summary['TTR_GC'] = event_drawdown(ln_Returns[(ln_Returns.index >= GC_date)], 10)
summary['DD_Covid'], summary['TTR_Covid'] = event_drawdown(ln_Returns[(ln_Returns.index >= Covid_date)], 20)
summary['DD_Max'], summary['TTR_Max'] = max_drawdown(ln_Returns[(ln_Returns.index >= Max_date)])


# Beta
for i in range(len(summary_rows)):
    summary.loc[summary_rows[i], 'Beta'] = round(np.cov(ln_Returns.iloc[-252:, i], ln_Returns['Ibovespa'].iloc[-252:])[0][1] / np.var(ln_Returns['Ibovespa'].iloc[-252:]),2)


# Delete indicators that have insuficient track record
summary['Ret_12M'] = np.where(ln_Returns[(ln_Returns.index >= date_12M)].iloc[0:3].isna().sum()>=2, 0, summary['Ret_12M'])
summary['Ret_24M'] = np.where(ln_Returns[(ln_Returns.index >= date_24M)].iloc[0:3].isna().sum()>=2, 0, summary['Ret_24M'])
summary['Ret_36M'] = np.where(ln_Returns[(ln_Returns.index >= date_36M)].iloc[0:3].isna().sum()>=2, 0, summary['Ret_36M'])

summary['Vol_12M'] = np.where(ln_Returns[(ln_Returns.index >= date_12M)].iloc[0:3].isna().sum()>=2, 0, summary['Vol_12M'])
summary['Vol_24M'] = np.where(ln_Returns[(ln_Returns.index >= date_24M)].iloc[0:3].isna().sum()>=2, 0, summary['Vol_24M'])

summary['Sharpe_12M'] = np.where(ln_Returns[(ln_Returns.index >= date_12M)].iloc[0:3].isna().sum()>=2, 0, summary['Sharpe_12M'])
summary['Sharpe_24M'] = np.where(ln_Returns[(ln_Returns.index >= date_24M)].iloc[0:3].isna().sum()>=2, 0, summary['Sharpe_24M'])
summary['Sharpe_36M'] = np.where(ln_Returns[(ln_Returns.index >= date_36M)].iloc[0:3].isna().sum()>=2, 0, summary['Sharpe_36M'])

#summary['DD_JD'] = np.where(ln_Returns[(ln_Returns.index >= JD_date)].iloc[0:3].isna().sum()>=2, 0, summary['DD_JD'])
#summary['DD_GC'] = np.where(ln_Returns[(ln_Returns.index >= GC_date)].iloc[0:3].isna().sum()>=2, 0, summary['DD_GC'])
summary['DD_Covid'] = np.where(ln_Returns[(ln_Returns.index >= Covid_date)].iloc[0:3].isna().sum()>=2, 0, summary['DD_Covid'])
#summary['TTR_JD'] = np.where(ln_Returns[(ln_Returns.index >= JD_date)].iloc[0:3].isna().sum()>=2, 0, summary['TTR_JD'])
#summary['TTR_GC'] = np.where(ln_Returns[(ln_Returns.index >= GC_date)].iloc[0:3].isna().sum()>=2, 0, summary['TTR_GC'])
summary['TTR_Covid'] = np.where(ln_Returns[(ln_Returns.index >= Covid_date)].iloc[0:3].isna().sum()>=2, 0, summary['TTR_Covid'])

summary['Beta'] = np.where(ln_Returns.iloc[-252:].isna().sum()>=2, 0, summary['Beta'])


# Set cdi statistics zero
summary.iloc[0,11:] = "" 
summary.iloc[0,5:9] = "" 

''''X) EXPORT TO EXCEL -----------------------------------------------------------------------------------------'''
                                
# Create workbook object (try to opeen an existing one, if it doesn`t exist, create one)
try:
    wb = openpyxl.load_workbook(excel_path)
except:
    wb = openpyxl.Workbook()
    wb.save(excel_path)
    wb = openpyxl.load_workbook(excel_path)

        
# Print Log Returns:
output_sheet = 'Log Returns'        

if not output_sheet in wb.sheetnames:
    worksheet = wb.create_sheet(output_sheet)
else: worksheet = wb[output_sheet]

for col in range(1, worksheet.max_column): # Delete old data
    for row in range(1, worksheet.max_row):  
        worksheet.cell(row=row, column=col).value = None

data = ln_Returns
rows = dataframe_to_rows(data, index=True)
for r_idx, row in enumerate(rows):
    for c_idx, value in enumerate(row):
        worksheet.cell(row=r_idx+2, column=c_idx+1, value=value)

        
# Print Summary Table:
output_sheet = 'Ranking'

if not output_sheet in wb.sheetnames:
    worksheet = wb.create_sheet(output_sheet)
else: worksheet = wb[output_sheet]

for col in range(2, worksheet.max_column): # Delete old data
    for row in range(4, worksheet.max_row):  
        worksheet.cell(row=row, column=col).value = None
     
data = summary
rows = dataframe_to_rows(data, index=True,  header=False)
for r_idx, row in enumerate(rows):
    for c_idx, value in enumerate(row):
        worksheet.cell(row=r_idx+4, column=c_idx+2, value=value)

        
wb.save(excel_path)
        
#data_inicial

# codigo 1
# Pegar nome dos fundos do excel
# Puxar cotas da base de dados
# Realizar os calculos dos indicadores de performance, matrizes de correlação e betas para varias janelas
# exportar para o excel esses valores

# codigo 2
# pegar os fundos do excel
# limpar os dados historicos antigos no excel
# pegar os retornos historicos dos fundos selecionados
# exportar para o excel os retornos diários históricos acumulados para determinadas janelas para colocar em gráfico
